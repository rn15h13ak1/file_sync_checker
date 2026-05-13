"""レポート生成 (Excel / HTML)。"""
from __future__ import annotations

import html
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook.workbook import Workbook as WB

from comparator import (
    STATUS_HASH_MISMATCH,
    STATUS_OK,
    STATUS_PARTIAL_MISSING,
    STATUS_PARTIAL_PRESENT,
    ComparisonResult,
    FileRow,
)
from scanner import FileEntry, ScanResult
from utils import human_bytes


MISSING_PLACEHOLDER = "－"
DATETIME_FMT = "%Y-%m-%d %H:%M:%S"


# === 配色 ===
FILL_OK = PatternFill("solid", fgColor="C6EFCE")           # 緑
FILL_HASH_MISMATCH = PatternFill("solid", fgColor="FFC7CE") # 赤
FILL_MISSING = PatternFill("solid", fgColor="FFD8A8")      # オレンジ
FILL_PARTIAL = PatternFill("solid", fgColor="FFE699")      # 黄
FILL_GRAY = PatternFill("solid", fgColor="D9D9D9")
FILL_HEADER = PatternFill("solid", fgColor="305496")
FONT_HEADER = Font(bold=True, color="FFFFFF")


@dataclass
class ReportContext:
    started_at: datetime
    finished_at: datetime
    config_path: Path
    scans: List[ScanResult]
    comparison: ComparisonResult


# ============================================================
# Excel
# ============================================================
def write_excel(ctx: ReportContext, out_path: Path) -> Path:
    wb: WB = Workbook()
    # デフォルトで作成される空シートを削除
    default_ws = wb.active
    wb.remove(default_ws)

    _excel_summary(wb, ctx)
    _excel_all_files(wb, ctx)
    _excel_subset(wb, ctx, "ハッシュ不一致", ctx.comparison.hash_mismatches)
    _excel_subset(wb, ctx, "ファイル欠落", ctx.comparison.missing_files)
    _excel_subset(wb, ctx, "余分なファイル", ctx.comparison.extra_files)
    _excel_dir_diff(wb, ctx)
    _excel_errors(wb, ctx)

    wb.save(out_path)
    return out_path


def _set_header_row(ws, headers: List[str]) -> None:
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _excel_summary(wb: WB, ctx: ReportContext) -> None:
    ws = wb.create_sheet("サマリー")
    elapsed = (ctx.finished_at - ctx.started_at).total_seconds()

    rows: List[List[object]] = [
        ["項目", "値"],
        ["スキャン開始", ctx.started_at.strftime(DATETIME_FMT)],
        ["スキャン終了", ctx.finished_at.strftime(DATETIME_FMT)],
        ["所要時間 (秒)", f"{elapsed:.2f}"],
        ["設定ファイル", str(ctx.config_path)],
        ["拠点数", len(ctx.scans)],
        [],
        ["拠点", "ルートパス", "ファイル数", "総サイズ", "エラー数"],
    ]
    for s in ctx.scans:
        total_size = sum(f.size for f in s.files.values())
        rows.append([
            s.location_name,
            str(s.root),
            len(s.files),
            human_bytes(total_size),
            len(s.errors),
        ])

    rows.append([])
    rows.append(["差分種別", "件数"])
    rows.append(["ハッシュ不一致", len(ctx.comparison.hash_mismatches)])
    rows.append(["ファイル欠落 (一部拠点になし)", len(ctx.comparison.missing_files)])
    rows.append(["余分なファイル (一部拠点のみ)", len(ctx.comparison.extra_files)])
    rows.append(["フォルダ構造差分", len(ctx.comparison.dir_diffs)])

    for r_idx, row in enumerate(rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1 or (isinstance(val, str) and val in {"拠点", "差分種別", "項目"}):
                # ヘッダ行の見た目
                if c_idx <= len(row):
                    cell.font = Font(bold=True)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10


def _file_columns_for(location_names: List[str]) -> List[str]:
    cols = ["No.", "相対パス", "ファイル名", "拡張子", "状態"]
    for name in location_names:
        cols.extend([f"{name}: サイズ", f"{name}: ハッシュ", f"{name}: 更新日時"])
    return cols


def _status_fill(status: str) -> Optional[PatternFill]:
    return {
        STATUS_OK: FILL_OK,
        STATUS_HASH_MISMATCH: FILL_HASH_MISMATCH,
        STATUS_PARTIAL_MISSING: FILL_MISSING,
        STATUS_PARTIAL_PRESENT: FILL_PARTIAL,
    }.get(status)


def _write_file_row(
    ws,
    row_idx: int,
    no: int,
    file_row: FileRow,
    location_names: List[str],
    distinct_hashes: int,
) -> None:
    rel = file_row.relpath
    name = rel.rsplit("/", 1)[-1]
    ext = Path(name).suffix

    ws.cell(row=row_idx, column=1, value=no)
    ws.cell(row=row_idx, column=2, value=rel)
    ws.cell(row=row_idx, column=3, value=name)
    ws.cell(row=row_idx, column=4, value=ext)

    status_cell = ws.cell(row=row_idx, column=5, value=file_row.status)
    fill = _status_fill(file_row.status)
    if fill is not None:
        status_cell.fill = fill

    col = 6
    for loc in location_names:
        entry: Optional[FileEntry] = file_row.entries.get(loc)
        if entry is None:
            for offset in range(3):
                c = ws.cell(row=row_idx, column=col + offset, value=MISSING_PLACEHOLDER)
                c.fill = FILL_GRAY
                c.alignment = Alignment(horizontal="center")
        else:
            ws.cell(row=row_idx, column=col, value=entry.size)
            hash_cell = ws.cell(row=row_idx, column=col + 1, value=entry.hash)
            ws.cell(row=row_idx, column=col + 2, value=entry.mtime.strftime(DATETIME_FMT))
            if file_row.status == STATUS_HASH_MISMATCH and distinct_hashes > 1:
                hash_cell.fill = FILL_HASH_MISMATCH
        col += 3


def _excel_all_files(wb: WB, ctx: ReportContext) -> None:
    ws = wb.create_sheet("全ファイル一覧")
    loc_names = ctx.comparison.location_names
    headers = _file_columns_for(loc_names)
    _set_header_row(ws, headers)

    for i, row in enumerate(ctx.comparison.all_files, 1):
        distinct = len({e.hash for e in row.entries.values() if e is not None})
        _write_file_row(ws, i + 1, i, row, loc_names, distinct)

    # 列幅
    ws.column_dimensions["A"].width = 6   # No.
    ws.column_dimensions["B"].width = 50  # 相対パス
    ws.column_dimensions["C"].width = 28  # ファイル名
    ws.column_dimensions["D"].width = 8   # 拡張子
    ws.column_dimensions["E"].width = 14  # 状態
    base = 6
    for _ in loc_names:
        ws.column_dimensions[get_column_letter(base)].width = 12      # サイズ
        ws.column_dimensions[get_column_letter(base + 1)].width = 16  # ハッシュ (短く表示, セル選択でフル値確認)
        ws.column_dimensions[get_column_letter(base + 2)].width = 20  # 更新日時
        base += 3

    # 固定: 1行目 + 相対パス列 (B列まで固定 = C列以降スクロール)
    ws.freeze_panes = "C2"

    # オートフィルタ
    ws.auto_filter.ref = ws.dimensions


def _excel_subset(wb: WB, ctx: ReportContext, sheet_name: str, rows: List[FileRow]) -> None:
    ws = wb.create_sheet(sheet_name)
    loc_names = ctx.comparison.location_names
    headers = _file_columns_for(loc_names)
    _set_header_row(ws, headers)

    if not rows:
        ws.cell(row=2, column=1, value="該当なし")
        return

    for i, row in enumerate(rows, 1):
        distinct = len({e.hash for e in row.entries.values() if e is not None})
        _write_file_row(ws, i + 1, i, row, loc_names, distinct)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 14
    base = 6
    for _ in loc_names:
        ws.column_dimensions[get_column_letter(base)].width = 12
        ws.column_dimensions[get_column_letter(base + 1)].width = 16
        ws.column_dimensions[get_column_letter(base + 2)].width = 20
        base += 3
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions


def _excel_dir_diff(wb: WB, ctx: ReportContext) -> None:
    ws = wb.create_sheet("フォルダ構造差分")
    loc_names = ctx.comparison.location_names
    headers = ["No.", "相対パス"] + loc_names
    _set_header_row(ws, headers)

    if not ctx.comparison.dir_diffs:
        ws.cell(row=2, column=1, value="該当なし")
        return

    for i, d in enumerate(ctx.comparison.dir_diffs, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=d.relpath)
        for j, loc in enumerate(loc_names):
            v = "○" if d.presence[loc] else MISSING_PLACEHOLDER
            cell = ws.cell(row=i + 1, column=3 + j, value=v)
            cell.alignment = Alignment(horizontal="center")
            if not d.presence[loc]:
                cell.fill = FILL_GRAY

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 60
    for j in range(len(loc_names)):
        ws.column_dimensions[get_column_letter(3 + j)].width = 12
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = ws.dimensions


def _excel_errors(wb: WB, ctx: ReportContext) -> None:
    ws = wb.create_sheet("エラー")
    _set_header_row(ws, ["No.", "拠点", "相対パス", "メッセージ"])

    rows: List[tuple] = []
    for s in ctx.scans:
        for e in s.errors:
            rows.append((s.location_name, e.relpath, e.message))

    if not rows:
        ws.cell(row=2, column=1, value="該当なし")
        return

    for i, (loc, rel, msg) in enumerate(rows, 1):
        ws.cell(row=i + 1, column=1, value=i)
        ws.cell(row=i + 1, column=2, value=loc)
        ws.cell(row=i + 1, column=3, value=rel)
        ws.cell(row=i + 1, column=4, value=msg)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 80
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ============================================================
# HTML
# ============================================================
def write_html(ctx: ReportContext, out_path: Path) -> Path:
    elapsed = (ctx.finished_at - ctx.started_at).total_seconds()
    loc_names = ctx.comparison.location_names

    summary_html = _html_summary_section(ctx, elapsed)
    all_files_html = _html_file_table("全ファイル一覧", ctx.comparison.all_files, loc_names)
    mismatch_html = _html_file_table("ハッシュ不一致", ctx.comparison.hash_mismatches, loc_names)
    missing_html = _html_file_table("ファイル欠落", ctx.comparison.missing_files, loc_names)
    extra_html = _html_file_table("余分なファイル", ctx.comparison.extra_files, loc_names)
    dir_html = _html_dir_diff(ctx.comparison.dir_diffs, loc_names)
    err_html = _html_errors(ctx.scans)

    title = f"File Sync Check Report - {ctx.started_at.strftime(DATETIME_FMT)}"
    body = f"""
<!DOCTYPE html>
<html lang="ja">
<head>
<meta charset="utf-8">
<title>{html.escape(title)}</title>
<style>
{_HTML_STYLE}
</style>
</head>
<body>
<header>
  <h1>ファイル同期チェック レポート</h1>
  <nav>
    <a href="#summary">サマリー</a>
    <a href="#all">全ファイル一覧</a>
    <a href="#mismatch">ハッシュ不一致</a>
    <a href="#missing">ファイル欠落</a>
    <a href="#extra">余分なファイル</a>
    <a href="#dirs">フォルダ構造差分</a>
    <a href="#errors">エラー</a>
  </nav>
</header>
<main>
  {summary_html}
  {all_files_html}
  {mismatch_html}
  {missing_html}
  {extra_html}
  {dir_html}
  {err_html}
</main>
</body>
</html>
"""
    out_path.write_text(body, encoding="utf-8")
    return out_path


_HTML_STYLE = """
* { box-sizing: border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", "Hiragino Sans",
       "Yu Gothic", sans-serif; margin: 0; color: #222; background: #fafafa; }
header { background: #305496; color: #fff; padding: 16px 24px; position: sticky; top: 0; z-index: 10; }
header h1 { margin: 0 0 8px; font-size: 1.3rem; }
header nav a { color: #fff; margin-right: 12px; text-decoration: none; font-size: 0.9rem; }
header nav a:hover { text-decoration: underline; }
main { padding: 24px; }
section { background: #fff; border-radius: 6px; padding: 16px; margin-bottom: 24px;
          box-shadow: 0 1px 3px rgba(0,0,0,0.08); }
section h2 { margin-top: 0; font-size: 1.1rem; border-bottom: 2px solid #305496; padding-bottom: 6px; }
table { border-collapse: collapse; width: 100%; font-size: 0.85rem; }
th, td { border: 1px solid #ddd; padding: 4px 8px; text-align: left; vertical-align: top;
         white-space: nowrap; overflow: hidden; text-overflow: ellipsis; max-width: 320px; }
th { background: #305496; color: #fff; position: sticky; top: 0; }
tr:nth-child(even) td { background: #f6f8fb; }
td.hash { font-family: ui-monospace, "SF Mono", Menlo, Consolas, monospace; font-size: 0.78rem; }
td.num { text-align: right; font-variant-numeric: tabular-nums; }
.status-ok { background: #c6efce !important; color: #006100; font-weight: bold; }
.status-mismatch { background: #ffc7ce !important; color: #9c0006; font-weight: bold; }
.status-missing { background: #ffd8a8 !important; color: #9c4500; font-weight: bold; }
.status-partial { background: #ffe699 !important; color: #7f6000; font-weight: bold; }
.cell-mismatch { background: #ffc7ce !important; }
.cell-missing { background: #d9d9d9 !important; color: #888; text-align: center; }
.summary-table td:first-child { font-weight: bold; width: 240px; background: #f0f3f8; }
.scroll-wrap { overflow-x: auto; max-width: 100%; }
.fixed-col-table th:nth-child(2), .fixed-col-table td:nth-child(2) {
  position: sticky; left: 0; background: inherit; z-index: 1;
}
.empty { color: #888; font-style: italic; padding: 8px; }
"""


def _status_class(status: str) -> str:
    return {
        STATUS_OK: "status-ok",
        STATUS_HASH_MISMATCH: "status-mismatch",
        STATUS_PARTIAL_MISSING: "status-missing",
        STATUS_PARTIAL_PRESENT: "status-partial",
    }.get(status, "")


def _section_id(title: str) -> str:
    return {
        "全ファイル一覧": "all",
        "ハッシュ不一致": "mismatch",
        "ファイル欠落": "missing",
        "余分なファイル": "extra",
    }.get(title, html.escape(title))


def _html_summary_section(ctx: ReportContext, elapsed: float) -> str:
    rows = [
        ("スキャン開始", ctx.started_at.strftime(DATETIME_FMT)),
        ("スキャン終了", ctx.finished_at.strftime(DATETIME_FMT)),
        ("所要時間 (秒)", f"{elapsed:.2f}"),
        ("設定ファイル", str(ctx.config_path)),
        ("拠点数", str(len(ctx.scans))),
    ]
    summary_kv = "".join(
        f"<tr><td>{html.escape(k)}</td><td>{html.escape(v)}</td></tr>" for k, v in rows
    )

    loc_rows = []
    for s in ctx.scans:
        total_size = sum(f.size for f in s.files.values())
        loc_rows.append(
            "<tr>"
            f"<td>{html.escape(s.location_name)}</td>"
            f"<td>{html.escape(str(s.root))}</td>"
            f"<td class='num'>{len(s.files):,}</td>"
            f"<td class='num'>{html.escape(human_bytes(total_size))}</td>"
            f"<td class='num'>{len(s.errors)}</td>"
            "</tr>"
        )

    diff_rows = [
        ("ハッシュ不一致", len(ctx.comparison.hash_mismatches)),
        ("ファイル欠落 (一部拠点になし)", len(ctx.comparison.missing_files)),
        ("余分なファイル (一部拠点のみ)", len(ctx.comparison.extra_files)),
        ("フォルダ構造差分", len(ctx.comparison.dir_diffs)),
    ]
    diff_html = "".join(
        f"<tr><td>{html.escape(k)}</td><td class='num'>{v:,}</td></tr>" for k, v in diff_rows
    )

    return f"""
<section id="summary">
  <h2>サマリー</h2>
  <table class="summary-table">{summary_kv}</table>
  <h3>拠点別</h3>
  <div class="scroll-wrap"><table>
    <thead><tr><th>拠点</th><th>ルートパス</th><th>ファイル数</th><th>総サイズ</th><th>エラー数</th></tr></thead>
    <tbody>{"".join(loc_rows)}</tbody>
  </table></div>
  <h3>差分件数</h3>
  <table class="summary-table">{diff_html}</table>
</section>
"""


def _html_file_table(title: str, rows: List[FileRow], loc_names: List[str]) -> str:
    sid = _section_id(title)
    if not rows:
        return f"""
<section id="{sid}">
  <h2>{html.escape(title)}</h2>
  <p class="empty">該当なし</p>
</section>
"""

    headers = ["No.", "相対パス", "ファイル名", "拡張子", "状態"]
    for n in loc_names:
        headers.extend([f"{n}: サイズ", f"{n}: ハッシュ", f"{n}: 更新日時"])

    head_html = "".join(f"<th>{html.escape(h)}</th>" for h in headers)

    body_lines: List[str] = []
    for i, row in enumerate(rows, 1):
        name = row.relpath.rsplit("/", 1)[-1]
        ext = Path(name).suffix
        status_cls = _status_class(row.status)
        cells: List[str] = [
            f"<td class='num'>{i}</td>",
            f"<td title='{html.escape(row.relpath)}'>{html.escape(row.relpath)}</td>",
            f"<td>{html.escape(name)}</td>",
            f"<td>{html.escape(ext)}</td>",
            f"<td class='{status_cls}'>{html.escape(row.status)}</td>",
        ]
        distinct_hashes = len({e.hash for e in row.entries.values() if e is not None})

        for loc in loc_names:
            entry = row.entries.get(loc)
            if entry is None:
                cells.append(f"<td class='cell-missing'>{MISSING_PLACEHOLDER}</td>")
                cells.append(f"<td class='cell-missing'>{MISSING_PLACEHOLDER}</td>")
                cells.append(f"<td class='cell-missing'>{MISSING_PLACEHOLDER}</td>")
            else:
                hash_cls = "hash"
                if row.status == STATUS_HASH_MISMATCH and distinct_hashes > 1:
                    hash_cls += " cell-mismatch"
                cells.append(f"<td class='num'>{entry.size:,}</td>")
                cells.append(
                    f"<td class='{hash_cls}' title='{html.escape(entry.hash)}'>"
                    f"{html.escape(entry.hash[:12])}…</td>"
                )
                cells.append(f"<td>{html.escape(entry.mtime.strftime(DATETIME_FMT))}</td>")
        body_lines.append("<tr>" + "".join(cells) + "</tr>")

    return f"""
<section id="{sid}">
  <h2>{html.escape(title)} ({len(rows):,} 件)</h2>
  <div class="scroll-wrap"><table class="fixed-col-table">
    <thead><tr>{head_html}</tr></thead>
    <tbody>{"".join(body_lines)}</tbody>
  </table></div>
</section>
"""


def _html_dir_diff(dir_diffs, loc_names: List[str]) -> str:
    if not dir_diffs:
        return """
<section id="dirs"><h2>フォルダ構造差分</h2><p class="empty">該当なし</p></section>
"""
    headers = ["No.", "相対パス"] + loc_names
    head_html = "".join(f"<th>{html.escape(h)}</th>" for h in headers)
    rows = []
    for i, d in enumerate(dir_diffs, 1):
        cells = [f"<td class='num'>{i}</td>", f"<td>{html.escape(d.relpath)}</td>"]
        for loc in loc_names:
            if d.presence[loc]:
                cells.append("<td style='text-align:center;'>○</td>")
            else:
                cells.append(f"<td class='cell-missing'>{MISSING_PLACEHOLDER}</td>")
        rows.append("<tr>" + "".join(cells) + "</tr>")
    return f"""
<section id="dirs">
  <h2>フォルダ構造差分 ({len(dir_diffs):,} 件)</h2>
  <div class="scroll-wrap"><table>
    <thead><tr>{head_html}</tr></thead>
    <tbody>{"".join(rows)}</tbody>
  </table></div>
</section>
"""


def _html_errors(scans: List[ScanResult]) -> str:
    rows = []
    n = 0
    for s in scans:
        for e in s.errors:
            n += 1
            rows.append(
                "<tr>"
                f"<td class='num'>{n}</td>"
                f"<td>{html.escape(s.location_name)}</td>"
                f"<td>{html.escape(e.relpath)}</td>"
                f"<td>{html.escape(e.message)}</td>"
                "</tr>"
            )
    if not rows:
        return """
<section id="errors"><h2>エラー</h2><p class="empty">該当なし</p></section>
"""
    return f"""
<section id="errors">
  <h2>エラー ({n:,} 件)</h2>
  <div class="scroll-wrap"><table>
    <thead><tr><th>No.</th><th>拠点</th><th>相対パス</th><th>メッセージ</th></tr></thead>
    <tbody>{"".join(rows)}</tbody>
  </table></div>
</section>
"""
